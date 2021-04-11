import threading
import win32com.client          ##################################
import time                     #                                #
import boto3                    #     last modify 2021.03.11     #
from random import *            #                                #
import telegram                 ##################################
from pykiwoom.kiwoom import *
import openpyxl

############################# AWS 연동 ########################

import boto3

dynamodb = boto3.resource(
"dynamodb",
aws_access_key_id="AKIA267WZH5NKOE3UHMK",
aws_secret_access_key="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",
region_name="us-east-2"
)

table = dynamodb.Table('StockTable-l3qhg25eiregjprodnybb5gs5e-staging')
dbIndex = 0 #dynamodb 항목 수
############################cybos 접근#########################
objCpCybos = win32com.client.Dispatch("CpUtil.CpCybos")
kiwoom = Kiwoom()
kiwoom.CommConnect(block=True)
print("블록킹 로그인 완료")
bConnect = objCpCybos.IsConnect
if (bConnect == 0):
    print("PLUS가 정상적으로 연결되지 않음. ")
    exit()
    
objStockWeek = win32com.client.Dispatch("DsCbo1.StockWeek")    
g_objCodeMgr = win32com.client.Dispatch('CpUtil.CpCodeMgr')
objStockMst = win32com.client.Dispatch("DsCbo1.StockMst")
objStockChart = win32com.client.Dispatch("CpSysDib.StockChart")
objRq = win32com.client.Dispatch("CpSysDib.MarketEye")
################################################################

telgm_token = '1144360580:AAF3hIjZfhWfWgjnx8DY-vvt0dQwRdeqFLo'
bot = telegram.Bot(token = telgm_token)

codeList = g_objCodeMgr.GetStockListByMarket(1) # 
codeList2 = g_objCodeMgr.GetStockListByMarket(2) # 
allcodelist = codeList + codeList2

wb = openpyxl.load_workbook('3월.xlsx')

sheet = wb.active


print(dbIndex)

stockList = []
risingStockList = []
class Stock :
    def __init__(self, code) :
        self.code = code
        self.graph = []
        self.avgGraph = []
        self.avg3Graph = []
        self.avg20Graph = []
        self.rsiGraph = []
        self.isBought = False
        self.size = 0
        self.avgSize = 0
        self.avg3Size = 0
        self.avg20Size = 0
        self.buyPrice = 0
        self.buyAmount = 0
        self.au = 0
        self.ad = 0
        self.isUpdated = False
        self.stopBuy = 0
        self.buyDate = 0
        self.nowPercent = 0
        self.lastDayClose = 0
        self.isRising = True
        self.point = 0
        self.buyTime = ""
        self.codeNo = 0
        
    def push(self, num) :
        self.graph.append(num)
        self.size = self.size+1
        if(len(self.graph)>10) :
            self.pushAvgGraph()
        if(len(self.graph)>23) :
            self.pushAvg20Graph()
            self.pushRsiGraph()
            
    def checkStatus(self) :
        avgSize = self.avgSize
        if((max(self.avgGraph[avgSize-1], self.avgGraph[avgSize-8])-min(self.avgGraph[avgSize-1], self.avgGraph[avgSize-8]))/(min(self.avgGraph[avgSize-1], self.avgGraph[avgSize-8]))) < 0.004 :
            self.status = "normal"
        elif((self.avgGraph[avgSize-1] < self.avgGraph[avgSize-4])&(self.avgGraph[avgSize-1] < self.avgGraph[avgSize-2])) :
            if(self.status == "rising")|(self.status == "falling") :
                maxValue = self.avgGraph[avgSize-1]
                for i in range(10) :
                    if(maxValue < self.avgGraph[avgSize-2-i]) :
                        maxValue = self.avgGraph[avgSize-2-i]
                if(self.highPrice2 == -1) :
                    self.highPrice2 = maxValue
                else :
                    self.highPrice1 = self.highPrice2
                    self.highPrice2 = maxValue
            self.status = "falling"
        elif((self.avgGraph[avgSize-1] > self.avgGraph[avgSize-4])&(self.avgGraph[avgSize-1] > self.avgGraph[avgSize-2])) :
            if(self.status == "falling")|(self.status == "normal") :
                minValue = self.avgGraph[avgSize-1]
                for i in range(10) :
                    if(minValue > self.avgGraph[avgSize-2-i]) :
                        minValue = self.avgGraph[avgSize-2-i]
                if(self.lowPrice2 == -1) :
                    self.lowPrice2 = minValue
                else :
                    self.lowPrice1 = self.lowPrice2
                    self.lowPrice2 = minValue
            self.status = "rising"
            
    def pushAvg20Graph(self) :
        sum = 0
        for i in range(20) :
            sum = sum + self.graph[self.size-1-i]
        self.avg20Graph.append(sum/20)
        self.avg20Size = self.avg20Size+1
        
    def pushAvgGraph(self) :
        sum = 0
        for i in range(10) :
            sum = sum + self.graph[self.size-1-i]
        sum2 = 0
        for i in range(5) :
            sum2 = sum2 + self.graph[self.size-1-i]
        self.avgGraph.append(sum/10)
        self.avg3Graph.append(sum2/5)
        self.avg3Size = self.avg3Size+1
        self.avgSize = self.avgSize+1
        
    def check(self, num) :
        au = 0
        ad = 0
        if(num >= self.graph[self.size-2]) :
            au = (self.au*13+(num-self.graph[self.size-2]))/14
            ad = (self.ad*13)/14
        else :
            au = (self.au*13)/14
            ad = (self.ad*13+(self.graph[self.size-2]-num))/14

        return au/(au+ad)
    def pushRsiGraph(self) :
        u = []
        d = []
        for i in range(14) :
            if(self.graph[self.size-15+i] <= self.graph[self.size-14+i]) :
                u.append(self.graph[self.size-14+i]-self.graph[self.size-15+i])
            else :
                d.append(self.graph[self.size-15+i]-self.graph[self.size-14+i])

        if(self.au == 0) :
            if(len(u) != 0) :
                self.au = sum(u, 0.0) / 14
            else :
                self.au = 0
        else :
            if(self.graph[self.size-1] >= self.graph[self.size-2]) :
                self.au = (self.au*13+u[len(u)-1])/14
            else :
                self.au = (self.au*13)/14
                
        if(self.ad == 0) :
            if(len(d) != 0) :
                self.ad = sum(d, 0.0) / 14
            else :
                self.ad = 0
        else :
            if(self.graph[self.size-1] < self.graph[self.size-2]) :
                self.ad = (self.ad*13+d[len(d)-1])/14
            else :
                self.ad = (self.ad*13)/14
        try :
            self.rsiGraph.append(self.au/(self.au+self.ad))
        except :
            print("0으로나눔")


class AsyncTask:
    def __init__(self):
        pass
    
    def run(self) :
        print("running")
        dbIndex = int(sheet.cell(row=1, column=10).value)
        rsiSUM = 0
        for i in range(len(stockList)) :
            rsiSUM = rsiSUM + stockList[i].rsiGraph[len(stockList[i].rsiGraph)-1]
            stockList[i].lastDayClose = stockList[i].graph[len(stockList[i].graph)-1]
        rsiAVG = rsiSUM/len(stockList)
        rsiBenchmark = 0.15
        lowBenchmark = 0.97
        highBenchmark = 1.02
        print("rsi : " + str(rsiBenchmark) + " low : " + str(lowBenchmark) + " high : " + str(highBenchmark))
        date = 20210409
        moneyForOneStock = 200000
        benchmarkMinute = 3
        success = 0
        fail = 0
        count = 0
        while True :
            now = time.localtime()
            if now.tm_hour == 15 :
                if now.tm_min > 10 :
                    wb.save('3월.xlsx')
                    break
            rsiSUM = 0
            if rsiAVG > 0 :
                print("평균 rsi : " + str(rsiAVG))
            
            for i in range(len(stockList)) :
                now = time.localtime()
                if now.tm_hour == 9 :
                    if now.tm_min>19 & now.tm_min<22 :
                        for l in range(len(stockList)) :
                            if stockList[i].isRising == True :
                                if int(stockList[i].graph[len(stockList[i].graph)-1]) < int(stockList[i].avg20Graph[len(stockList[i].avg20Graph)-1])*1.02 :
                                    stockList[i].isRising = False
                close = -1
                try :
                    objStockChart.SetInputValue(0, stockList[i].code)
                    objStockChart.SetInputValue(1, ord('1'))
                    objStockChart.SetInputValue(2, date)
                    objStockChart.SetInputValue(3, date)
                    objStockChart.SetInputValue(4, 50) #
                    objStockChart.SetInputValue(5, (0, 1, 2, 3, 4, 5, 8, 9))
                    objStockChart.SetInputValue(6, ord('m'))
                    objStockChart.SetInputValue(9, ord('0'))
                    objStockChart.BlockRequest()
                    
                    objStockWeek.SetInputValue(0, stockList[i].code)
                    objStockWeek.BlockRequest()
                    close = objStockChart.GetDataValue(5, 0)
                except :
                    continue
                
                if (close == -1) :
                    continue
                if (now.tm_min%benchmarkMinute==0) & (stockList[i].isUpdated == False) :
                    stockList[i].push(close)
                    stockList[i].isUpdated = True
                if (now.tm_min%benchmarkMinute != 0) :
                    stockList[i].isUpdated = False
                if(len(stockList[i].rsiGraph)==0) :
                    #print(len(stockList[i].rsiGraph))
                    continue
                rsi = stockList[i].rsiGraph[len(stockList[i].rsiGraph)-1]
                rsiSUM = rsiSUM + rsi
                #print(g_objCodeMgr.CodeToName(stockList[i].code) + "\t" + rsi)
                
                
                if(stockList[i].isBought == True) :
                    isSell = False
                    if (now.tm_hour >= 15) :
                        isSell = True
                    if (close > stockList[i].buyPrice*highBenchmark) :
                        isSell = True
                    if (close < stockList[i].buyPrice*lowBenchmark) :
                        isSell = True
                    if isSell == True :
                        sellPrice = 0
                        if(close >= 1000)&(close<10000) :
                            sellPrice = close-15
                        if(close>=10000)&(close<100000) :
                            sellPrice = close-100
                        if(close>100000) :
                            sellPrice = close
                        stockList[i].isBought = False
                        print(g_objCodeMgr.CodeToName(stockList[i].code) + "\t" + str(sellPrice) + "\t매도")
                        cd = stockList[i].code
                        accounts = kiwoom.GetLoginInfo("ACCNO")
                        stock_account = accounts[0]
                        
                        kiwoom.SendOrder("RQ_1", "0101", stock_account, 2, str(cd[1:]), int(stockList[i].buyAmount), int(sellPrice), "00", "")
                        time.sleep(2)
                        if((float(close)*int(stockList[i].buyAmount))-(float(stockList[i].buyPrice)*int(stockList[i].buyAmount)) > 0) :
                            success = success+1
                        else :
                            fail = fail+1
                            
                        bot.sendMessage(chat_id = '1339775032', text=g_objCodeMgr.CodeToName(stockList[i].code) + "\n 매수 : " + str(stockList[i].buyPrice) + "\n 매도 : " + str(close) + "\n 성공 : " + str(success) + "\n 실패 : " + str(fail))
                        bot.sendMessage(chat_id = '1193408697', text=g_objCodeMgr.CodeToName(stockList[i].code) + "\n 매수 : " + str(stockList[i].buyPrice) + "\n 매도 : " + str(close) + "\n 성공 : " + str(success) + "\n 실패 : " + str(fail))
                        bot.sendMessage(chat_id = '1141981571', text=g_objCodeMgr.CodeToName(stockList[i].code) + "\n 매수 : " + str(stockList[i].buyPrice) + "\n 매도 : " + str(close) + "\n 성공 : " + str(success) + "\n 실패 : " + str(fail))
                        bot.sendMessage(chat_id = '1493196380', text=g_objCodeMgr.CodeToName(stockList[i].code) + "\n 매수 : " + str(stockList[i].buyPrice) + "\n 매도 : " + str(close) + "\n 성공 : " + str(success) + "\n 실패 : " + str(fail))
                        bot.sendMessage(chat_id = '1559254315', text=g_objCodeMgr.CodeToName(stockList[i].code) + "\n 매수 : " + str(stockList[i].buyPrice) + "\n 매도 : " + str(close) + "\n 성공 : " + str(success) + "\n 실패 : " + str(fail))
                        count = count-1
                        print("성공 : " + str(success))
                        print("실패 : " + str(fail))
                        sheet.cell(int(stockList[i].codeNo)+2,1, str(stockList[i].buyTime))
                        sheet.cell(int(stockList[i].codeNo)+2,2, str(rsiAVG))
                        sheet.cell(int(stockList[i].codeNo)+2,3, g_objCodeMgr.CodeToName(stockList[i].code))
                        sheet.cell(int(stockList[i].codeNo)+2,4, str(stockList[i].buyPrice))
                        sheet.cell(int(stockList[i].codeNo)+2,5, str(close))
                        sheet.cell(int(stockList[i].codeNo)+2,6, str(stockList[i].buyAmount))
                        sheet.cell(int(stockList[i].codeNo)+2,7, str(now.tm_mon) + "/" + str(now.tm_mday) + " " + str(now.tm_hour) + ":" + str(now.tm_min))
                        sheet.cell(int(stockList[i].codeNo)+2,8, str(float(stockList[i].buyPrice)*int(stockList[i].buyAmount)))
                        sheet.cell(int(stockList[i].codeNo)+2,9, str(float(close)*int(stockList[i].buyAmount)))
                    continue
                isBuy = False
                if now.tm_hour == 9 :
                    if now.tm_min < 20 :
                        if close < stockList[i].lastDayClose & now.tm_min > 14 :
                            if stockList[i].isRising == True :
                                stockList[i].isRising = False
                    else :
                        if stockList[i].isRising == True :
                            if stockList[i].point == 0 & close < int(stockList[i].avg20Graph[len(stockList[i].avg20Graph)-1])*1.0075 :
                                stockList[i].point = int(stockList[i].avg20Graph[len(stockList[i].avg20Graph)-1])
                                print(stockList[i].code)
                elif now.tm_hour < 11 :
                    if stockList[i].isRising == True :
                        if stockList[i].point == 0 & close < int(stockList[i].avg20Graph[len(stockList[i].avg20Graph)-1])*1.0075 :
                            stockList[i].point = int(stockList[i].avg20Graph[len(stockList[i].avg20Graph)-1])
                            print(stockList[i].code)
                elif now.tm_hour < 12 :
                    if stockList[i].point != 0 :
                        if close < stockList[i].point * 0.98 :
                            stockList[i].point = 0
                        if close > stockList[i].point * 1.015 :
                            #isBuy = True
                            print("상승")
                if (now.tm_hour == 14) :
                    if now.tm_min > 30 :
                        continue
                if now.tm_hour > 14 :
                    continue
                if count >= 4 :
                    continue
                if(rsi < rsiBenchmark) :
                    isBuy = True
                    print("rsi")
                    
                if isBuy :
                    stockList[i].buyTime = str(now.tm_mon) + "/" + str(now.tm_mday) + " " + str(now.tm_hour) + ":" + str(now.tm_min)
                    if(close >= 1200)&(close<10000) :
                        buyPrice = close + 25
                    elif(close >= 10000) & (close < 100000) :
                        buyPrice= close + 200
                    else :
                        continue
                    print(g_objCodeMgr.CodeToName(stockList[i].code) + "\t" + str(close) + "\t매수")
                    stockList[i].isBought = True
                    stockList[i].buyPrice = close
                    count = count+1
                    stockList[i].buyAmount = int(moneyForOneStock/close)
                    table.put_item(
                        Item={
                            'id' : str(dbIndex),
                            'code': stockList[i].code,
                            'name' : g_objCodeMgr.CodeToName(stockList[i].code),
                            'buyPrice': close,
                            'buyAmount': stockList[i].buyAmount,
                            'sellPrice' : 0,
                        }
                    )
                    try :
                        
                        cd = stockList[i].code
                        accounts = kiwoom.GetLoginInfo("ACCNO")
                        stock_account = accounts[0]

                        kiwoom.SendOrder("RQ_1", "0101", stock_account, 1, str(cd[1:]), int(stockList[i].buyAmount), buyPrice, "00", "")
                        time.sleep(2)
                        print("매수 요청")
                    except :
                        print("매수안댐")
                        continue
                    #putItem(str(dbIndex), stockList[i], g_objCodeMgr.CodeToName(stockList[i].code), close, stockList[i].buyAmount, 0)
                    stockList[i].codeNo = dbIndex
                    dbIndex = dbIndex+1
                    sheet.cell(1,10, str(dbIndex))
            rsiAVG = rsiSUM / len(stockList)        
    def collectData(self):
        print("collect data")
        date = 20210408
        benchmarkMinute = 3
        #print("test")
        for i in range(len(stockList)) :
            objStockChart.SetInputValue(0, stockList[i].code)
            objStockChart.SetInputValue(1, ord('1'))
            objStockChart.SetInputValue(2, date)
            objStockChart.SetInputValue(3, date)
            objStockChart.SetInputValue(4, 50) #
            objStockChart.SetInputValue(5, (0, 1, 2, 3, 4, 5, 8, 9))
            objStockChart.SetInputValue(6, ord('m'))
            objStockChart.SetInputValue(9, ord('0'))
            objStockChart.BlockRequest()

            for j in range(400, 0, -1) :
                #print("abc")
                close = -1
                try :
                    close = objStockChart.GetDataValue(5, j)
                except :
                    continue
                if close == -1 :
                    continue
                if j%benchmarkMinute == 0 :
                    stockList[i].push(close)
            
                        
def main() :

    loadList()
    print(len(stockList))
    at = AsyncTask()
    at.collectData()
    at.run()

def putItem(id, code, name, buyPrice, buyAmount, sellPrice) :
    table.put_item(
        Item={
            'id' : id,
            'code': code,
            'name' : name,
            'buyPrice': buyPrice,
            'buyAmount': buyAmount,
            'sellPrice' : sellPrice,
        }
    )

def loadList() :
    #load CheckList
    myFile = open('checkList.txt', 'r')
    while True :
        tmp = myFile.readline()[:-1]
        if tmp == 'X' :
            break
        stk = Stock(str(tmp))
        stockList.append(stk)
    myFile.close()
    
if __name__== "__main__" :
    main()
